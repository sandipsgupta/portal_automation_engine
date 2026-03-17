import csv
import json
import os
import sys
from datetime import datetime
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright

try:
    import openpyxl
except ImportError:
    openpyxl = None

# When packaged as .exe, resolve paths relative to the .exe location
# When running as script, resolve relative to main.py
_base = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) \
        else os.path.dirname(os.path.abspath(__file__))

_screenshots_dir = os.path.join(_base, "screenshots")
os.makedirs(_screenshots_dir, exist_ok=True)

# Load .env from the same folder as the .exe / script
load_dotenv(os.path.join(_base, ".env"))

USERNAME = os.getenv("PORTAL_USERNAME")
PASSWORD = os.getenv("PORTAL_PASSWORD")


# ──────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────
def load_config():
    config_path = os.path.join(_base, "config.json")
    with open(config_path) as f:
        return json.load(f)


# ──────────────────────────────────────────────────────────────────────
# CSV READER
# ──────────────────────────────────────────────────────────────────────

# Sheet column header → portal Payment Mode dropdown label
PAYMENT_MODE_MAP = {
    "Cash":          "Cash",
    "Cheque Amount": "Cheque",
    "UPI Amount":    "UPI",
}
# Credit Amount rows skipped — they use Add Credit Note, not Add Collections

# Portal table column header → payment mode (used for duplicate detection)
PORTAL_COL_MAP = {
    "Cash":   "Cash",
    "Cheque": "Cheque",
    "UPI":    "UPI",
    "NEFT":   "NEFT",
}


def load_rows_from_csv(csv_path: str):
    """
    Read invoice rows from CSV (.csv) or Excel (.xlsx/.xls).
    Columns used: Bill Number, Bill Date, Cash, Cheque Amount, UPI Amount,
                  Credit Amount

    Returns:
        rows     — list of payment entry dicts to process
        run_date — the Bill Date from the sheet (DD-MM-YYYY),
                   used to auto-set the portal date filter

    Rules:
      - Credit-only rows are SKIPPED (separate Add Credit Note flow)
      - Split payments (Cash + UPI on same row) produce TWO entries
      - Rows with no payable amount are skipped with a warning
    """
    rows     = []
    skipped  = []
    run_date = None

    # ── Load rows from CSV or Excel ───────────────────────────────────
    ext = str(csv_path).lower()
    if ext.endswith(".xlsx") or ext.endswith(".xls"):
        if openpyxl is None:
            print("⚠️  openpyxl not available — cannot read .xlsx files")
            raw_rows = []
        else:
            wb   = openpyxl.load_workbook(csv_path, data_only=True)
            ws   = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            raw_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_rows.append(
                    {headers[i]: (str(v).strip() if v is not None else "")
                     for i, v in enumerate(row)}
                )
    else:
        # CSV (default)
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            raw_rows = list(csv.DictReader(f))

    # ── Process rows ──────────────────────────────────────────────────
    for raw in raw_rows:
        invoice_no = raw.get("Bill Number", "").strip()
        date_raw   = raw.get("Bill Date",   "").strip()

        if not invoice_no:
            continue

        # Capture the date from the first real data row
        if run_date is None and date_raw:
            run_date = date_raw

        # ── Resolve dates ─────────────────────────────────────────────
        # Bill Date is informational only — never used as a payment date.
        # DELIVERY DATE is used for Cash and UPI.
        # CHEQUE DATE is used for Cheque, falls back to DELIVERY DATE.
        bill_year       = _parse_date(date_raw).year if date_raw else datetime.now().year
        delivery_raw    = raw.get("DELIVERY DATE",  "").strip()
        cheque_date_raw = raw.get("CHEQUE DATE",    "").strip()
        cheque_no_raw   = raw.get("CHEQUE NUMBER",  "").strip()

        delivery_date = (
            _parse_date(delivery_raw, ref_year=bill_year).strftime("%d-%m-%Y")
            if delivery_raw else ""
        )
        cheque_date = (
            _parse_date(cheque_date_raw, ref_year=bill_year).strftime("%d-%m-%Y")
            if cheque_date_raw else delivery_date
        )

        # Build one entry per non-empty payment column
        payments = []
        for col, mode in PAYMENT_MODE_MAP.items():
            val = raw.get(col, "").strip()
            if val:
                payment_date = cheque_date if mode == "Cheque" else delivery_date
                if not payment_date:
                    skipped.append(
                        f"  ⚠️  {invoice_no} [{mode}] — no payment date found, skipped"
                    )
                    continue
                payments.append({
                    "invoice_no":   invoice_no,
                    "payment_mode": mode,
                    "cheque_no":    cheque_no_raw if mode == "Cheque" else "",
                    "amount":       val,
                    "date":         payment_date,
                })

        credit = raw.get("Credit Amount", "").strip()
        if not payments and credit:
            skipped.append(
                f"  ⏭  {invoice_no} — Credit only (Rs.{credit}), skipped"
            )
            continue

        if not payments:
            skipped.append(
                f"  ⚠️  {invoice_no} — no payable amount found, skipped"
            )
            continue

        rows.extend(payments)

    if skipped:
        print("\nSkipped rows:")
        for s in skipped:
            print(s)

    return rows, run_date


# ──────────────────────────────────────────────────────────────────────
# DATE HELPERS
# ──────────────────────────────────────────────────────────────────────
def _parse_date(date_str, ref_year: int = None) -> datetime:
    """
    Parse any supported date string → datetime object.
    ref_year: fallback year used for short formats like '22-Aug' (no year).
    Also accepts datetime objects directly (as returned by openpyxl).
    """
    if isinstance(date_str, datetime):
        return date_str
    date_str = str(date_str).strip()
    for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue
    # Short format: DD-Mon (no year) — e.g. '22-Aug' from DELIVERY DATE column
    try:
        dt = datetime.strptime(date_str.strip(), "%d-%b")
        return dt.replace(year=ref_year or datetime.now().year)
    except ValueError:
        pass
    raise ValueError(
        f"Unrecognised date format: '{date_str}'. "
        "Supported: DD-Mon-YYYY, DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY, DD-Mon"
    )


def to_portal_filter_date(date_str: str) -> str:
    """
    Format for the portal's Invoice From/To Date filter inputs.
    The filter fields expect DD-Mon-YYYY (e.g. '27-Feb-2026').
    """
    return _parse_date(date_str).strftime("%d-%b-%Y")


def to_slds_date(date_str: str) -> str:
    """
    Format for the modal's payDate text input (SLDS date picker).
    Salesforce en-US locale expects MM/DD/YYYY.
    """
    return _parse_date(date_str).strftime("%m/%d/%Y")


# ──────────────────────────────────────────────────────────────────────
# ENGINE
# ──────────────────────────────────────────────────────────────────────
class RetailerPortalEngine:

    def __init__(self, config):
        self.config     = config
        self.playwright = None
        self.browser    = None
        self.context    = None
        self.page       = None

    # ── LAUNCH & LOGIN ─────────────────────────────────────────────────
    def launch(self):
        print("Launching browser...")
        # When running as a PyInstaller .exe, the bundled Playwright
        # looks in the temp extraction folder for browsers — they're not there.
        # Point it to the user's actual installed browsers location instead.
        if getattr(sys, 'frozen', False):
            browsers_path = os.path.join(
                os.environ.get("LOCALAPPDATA", ""),
                "ms-playwright"
            )
            os.environ["PLAYWRIGHT_BROWSERS_PATH"] = browsers_path
            print(f"  Browser path: {browsers_path}")
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=self.config.get("headless", False),
            args=["--ignore-certificate-errors"]
        )
        self.context = self.browser.new_context(ignore_https_errors=True)
        self.page    = self.context.new_page()

    def login(self):
        print("Logging in...")
        self.page.goto(self.config["portal_url"])
        self.page.get_by_placeholder("Username").fill(USERNAME)
        self.page.get_by_placeholder("Password").fill(PASSWORD)
        self.page.get_by_role("button", name="Log in").click()
        try:
            self.page.wait_for_load_state("domcontentloaded", timeout=30000)
        except Exception:
            pass
        self.page.wait_for_timeout(5000)
        print(f"Login complete. URL: {self.page.url}")

    # ── UTILITIES ──────────────────────────────────────────────────────
    def _wait_mask_gone(self, timeout=15000):
        try:
            self.page.locator("div.mask").wait_for(state="hidden", timeout=timeout)
        except Exception:
            pass

    def _clear_and_fill(self, locator, value: str, label: str = "field"):
        """Select-all then fill — works for plain text inputs."""
        locator.click(click_count=3)
        locator.fill(str(value))
        print(f"  ✅ {label}: {value}")

    def _fill_slds_date(self, locator, slds_date: str, label: str = "Date"):
        """
        Type into an SLDS text-type date picker (input[name='payDate']).
        Must use keyboard.type() — fill() bypasses LWC change events
        and leaves the Add Payment button disabled.
        Format: MM/DD/YYYY
        """
        locator.click()
        self.page.wait_for_timeout(200)
        locator.press("Control+a")
        self.page.wait_for_timeout(100)
        self.page.keyboard.type(slds_date, delay=80)
        self.page.wait_for_timeout(200)
        locator.press("Tab")
        self.page.wait_for_timeout(400)
        print(f"  ✅ {label}: {slds_date}")

    def _safe_mouse_click(self, locator, label="element"):
        """Click via bounding box — reliable in LWC Shadow DOM."""
        locator.scroll_into_view_if_needed()
        self.page.wait_for_timeout(200)
        box = locator.bounding_box()
        if box:
            self.page.mouse.click(
                box['x'] + box['width']  / 2,
                box['y'] + box['height'] / 2
            )
            print(f"  ✅ Clicked: {label}")
        else:
            locator.click(force=True)
            print(f"  ✅ Force-clicked (no bbox): {label}")

    # ── DUPLICATE DETECTION ────────────────────────────────────────────
    def _already_entered(self, target_row, payment_mode: str) -> bool:
        """
        Check if this payment mode is already filled in the table row.

        The portal table has columns: Cash | Cheque | UPI | NEFT …
        If the cell for the matching column already has a non-zero value,
        the entry was previously saved — skip it to avoid duplication.

        Works by reading the header row to find the column index,
        then reading that cell from the target row.
        """
        # Map payment mode → column header text in the portal table
        col_header_map = {
            "Cash":   "Cash",
            "Cheque": "Cheque",
            "UPI":    "UPI",
            "NEFT":   "NEFT",
        }
        col_header = col_header_map.get(payment_mode)
        if not col_header:
            return False   # unknown mode — don't skip, attempt entry

        try:
            # Find column index from header row
            header_cells = self.page.locator("table thead tr th")
            col_index = -1
            for i in range(header_cells.count()):
                if col_header.lower() in header_cells.nth(i).inner_text().lower():
                    col_index = i
                    break

            if col_index == -1:
                return False   # column not found — don't skip

            # Read that cell from the target row
            cell = target_row.locator("td").nth(col_index)
            cell_text = cell.inner_text().strip()

            # Non-zero means already entered
            try:
                if float(cell_text.replace(",", "")) != 0:
                    return True
            except ValueError:
                pass   # not a number (e.g. blank) — treat as not entered

        except Exception as e:
            print(f"  ⚠️  Duplicate check failed: {e} — proceeding anyway")

        return False

    # ── DIAGNOSTICS ────────────────────────────────────────────────────
    def debug_modal_fields(self, modal, tag=""):
        label = f" [{tag}]" if tag else ""
        print(f"\n=== MODAL FIELD DIAGNOSTICS{label} ===")
        inputs  = modal.locator("input")
        buttons = modal.locator("button")
        combos  = modal.locator("lightning-combobox, [role='combobox']")

        print(f"  Inputs ({inputs.count()}):")
        for i in range(inputs.count()):
            inp = inputs.nth(i)
            try:
                print(f"    [{i}] type={inp.get_attribute('type')!r}  "
                      f"name={inp.get_attribute('name')!r}  "
                      f"placeholder={inp.get_attribute('placeholder')!r}")
            except Exception:
                pass

        print(f"  Comboboxes ({combos.count()}):")
        for i in range(combos.count()):
            c = combos.nth(i)
            try:
                print(f"    [{i}] label={c.get_attribute('label')!r}  "
                      f"aria-label={c.get_attribute('aria-label')!r}")
            except Exception:
                pass

        print(f"  Buttons ({buttons.count()}):")
        for i in range(buttons.count()):
            b = buttons.nth(i)
            try:
                print(f"    [{i}] text={b.inner_text().strip()!r:25}  "
                      f"name={b.get_attribute('name')!r}  "
                      f"title={b.get_attribute('title')!r}")
            except Exception:
                pass
        print("=" * 40)

    # ── ROWS PER PAGE ──────────────────────────────────────────────────
    def _set_max_rows_per_page(self):
        """
        Find the rows-per-page selector and set it to its highest option
        so all invoices for the day appear in a single table view.

        Salesforce Experience Cloud portals typically render this as a
        <select> or lightning-combobox near the table footer.
        Fails silently with a warning so it never blocks a run.
        """
        try:
            # Common selector patterns for Salesforce page-size controls
            selector = (
                "select[name='pageSize'], "
                "select[name='rows'], "
                "select.slds-select"
            )
            ps = self.page.locator(selector).first
            if ps.count() == 0:
                ps = self.page.locator("select").filter(
                    has_text=lambda t: any(d in t for d in ["10", "25", "50"])
                ).first

            ps.wait_for(state="visible", timeout=5000)

            # Collect all <option> values and pick the largest numeric one
            options = ps.locator("option").all()
            best_val = None
            best_num = -1
            for opt in options:
                val = opt.get_attribute("value") or opt.inner_text().strip()
                try:
                    n = int(val)
                    if n > best_num:
                        best_num = n
                        best_val = val
                except ValueError:
                    pass

            if best_val:
                ps.select_option(best_val)
                self.page.wait_for_timeout(2000)
                self._wait_mask_gone()
                print(f"  ✅ Rows per page set to {best_val}")
            else:
                print("  ⚠️  Rows-per-page selector found but no numeric options detected.")

        except Exception as e:
            print(f"  ⚠️  Could not set rows per page: {e} — pagination may apply.")

    # ── NAVIGATE + SEARCH ──────────────────────────────────────────────
    def navigate_to_settlement_page(self, run_date: str):
        """
        run_date: DD-MM-YYYY string auto-read from the CSV.
        Converted to DD-Mon-YYYY for the portal filter inputs.
        Same date used for both from_date and to_date (client preps
        one day's sheet at a time).
        """
        portal_date = to_portal_filter_date(run_date)
        print(f"Navigating to settlement page (date: {portal_date})...")

        self.page.goto(
            self.config["settlement_url"],
            wait_until="domcontentloaded"
        )
        self.page.locator("label", has_text="Invoice From Date").first.wait_for(timeout=30000)
        self._wait_mask_gone()
        self.page.wait_for_timeout(500)
        print("Settlement page loaded.")

        # Fill date filter — same date for from and to
        self.page.locator("input[name='invoiceFromDate']").click()
        self.page.locator("input[name='invoiceFromDate']").fill(portal_date)
        self.page.keyboard.press("Tab")

        self.page.locator("input[name='invoiceToDate']").click()
        self.page.locator("input[name='invoiceToDate']").fill(portal_date)
        self.page.keyboard.press("Tab")

        self.page.wait_for_timeout(500)
        self._wait_mask_gone()

        # (A) Top purple Search
        from_label = self.page.locator("label", has_text="Invoice From Date").first
        to_label   = self.page.locator("label", has_text="Invoice To Date").first
        date_area  = self.page.locator("div").filter(has=from_label).filter(has=to_label).first
        top_search = date_area.get_by_role("button", name="Search", exact=True).first
        top_search.scroll_into_view_if_needed()
        self._wait_mask_gone()
        top_search.click()
        print("Top Search clicked.")
        self.page.wait_for_timeout(2000)
        self._wait_mask_gone()

        # (B) Bottom Search — loads results table
        all_search   = self.page.get_by_role("button", name="Search", exact=True)
        lower_search = all_search.nth(all_search.count() - 1)
        lower_search.scroll_into_view_if_needed()
        self._wait_mask_gone()
        try:
            lower_search.click(timeout=10000)
        except Exception:
            self._wait_mask_gone()
            lower_search.click(force=True, timeout=10000)
        print("Bottom Search clicked.")

        # Wait longer — LWC re-renders table after API response
        self.page.wait_for_timeout(6000)
        self._wait_mask_gone()
        self.page.wait_for_timeout(2000)
        self._wait_mask_gone()

        self._set_max_rows_per_page()

        _ss = os.path.join(_screenshots_dir, "01_after_search.png")
        self.page.screenshot(path=_ss, full_page=True)
        print(f"Search complete → {_ss}")

    # ── FILL ONE ROW ───────────────────────────────────────────────────
    def fill_one_row(self, invoice_no, payment_mode, cheque_no, amount, date):
        print(f"\n{'─'*55}")
        print(f"Processing: {invoice_no}  [{payment_mode}  Rs.{amount}]")

        slds_date = to_slds_date(date)
        print(f"  Date: {date!r} → {slds_date!r} (MM/DD/YYYY for portal)")

        debug = self.config.get("debug", False)

        # ── 1. Find the matching table row ────────────────────────────
        # wait_for_selector() uses CSS and cannot pierce LWC Shadow DOM.
        # Use Playwright locator.wait_for() instead — shadow-DOM aware.
        rows = self.page.locator("table tbody tr")
        rows.first.wait_for(state="visible", timeout=30000)
        row_count = rows.count()
        print(f"  Rows in table: {row_count}")

        target_row = None
        for i in range(row_count):
            row = rows.nth(i)
            if invoice_no in row.inner_text():
                target_row = row
                print(f"  ✅ Matched row {i}")
                break

        if target_row is None:
            print(f"  ❌ Invoice {invoice_no} not found in table. Skipping.")
            return

        # ── 2. Duplicate check ────────────────────────────────────────
        # Before opening the modal, verify this payment mode hasn't
        # already been entered in a previous run or earlier this session.
        if self._already_entered(target_row, payment_mode):
            print(f"  ⏭  {payment_mode} already has a value for {invoice_no} — skipping to avoid duplicate.")
            return

        target_row.scroll_into_view_if_needed()
        self.page.wait_for_timeout(800)

        # ── 3. Click row-level "Add Collections" (+) button ───────────
        row_add_btn = target_row.locator("button").first
        self._safe_mouse_click(row_add_btn, label="row + (Add Collections)")
        self.page.wait_for_timeout(2000)
        self._wait_mask_gone()

        # ── 4. Wait for modal ─────────────────────────────────────────
        modal = self.page.locator("div.slds-modal__container")
        modal.wait_for(state="visible", timeout=15000)
        self.page.wait_for_timeout(600)
        print("  ✅ Modal opened.")

        if debug:
            self.debug_modal_fields(modal, tag="initial")

        # ── 5. Payment Mode ───────────────────────────────────────────
        print(f"  Selecting payment mode: {payment_mode!r}")
        self._wait_mask_gone()   # clear any overlay before opening dropdown
        modal.locator("button[name='paymentMode']").click()
        self.page.wait_for_timeout(800)
        self._wait_mask_gone()   # clear mask that may close the dropdown

        # filter(has_text) is more robust than exact ARIA name matching in
        # Salesforce LWC where options may have nested spans / whitespace.
        # _safe_mouse_click uses bounding box — reliable in Shadow DOM.
        option = self.page.locator("[role='option']").filter(has_text=payment_mode).first
        option.wait_for(state="visible", timeout=10000)
        self._safe_mouse_click(option, label=f"option:{payment_mode}")
        self.page.wait_for_timeout(600)
        print(f"  ✅ Payment Mode: {payment_mode}")

        # ── 6. Cheque No (appears dynamically after selecting Cheque) ──
        if cheque_no:
            print(f"  Waiting for Cheque No field...")
            cheque_input = modal.locator(
                "input[name='chequeNo'], "
                "input[name='cheque_no'], "
                "input[name='chequeno'], "
                "input[placeholder*='heque']"
            ).first
            try:
                cheque_input.wait_for(state="visible", timeout=5000)
                self._clear_and_fill(cheque_input, cheque_no, "Cheque No")
            except Exception:
                print("  ⚠️  Cheque field not found by name — trying nth(2)...")
                self._clear_and_fill(
                    modal.locator("input").nth(2), cheque_no, "Cheque No (fallback)"
                )

        # ── 7. Amount ─────────────────────────────────────────────────
        self._clear_and_fill(
            modal.locator("input[name='amount']"), amount, "Amount"
        )

        # ── 8. Date (SLDS text date picker — must use keyboard.type) ──
        self._fill_slds_date(
            modal.locator("input[name='payDate']"),
            slds_date,
            label="Date"
        )

        # ── 8b. Transaction ID — required field, default "1" ──────────
        txn_input = modal.locator(
            "input[name='transactionId'], "
            "input[name='transaction_id'], "
            "input[name='txnId'], "
            "input[name='txn_id'], "
            "input[placeholder*='ransaction']"
        ).first
        try:
            if txn_input.is_visible(timeout=3000):
                self._clear_and_fill(txn_input, "1", "Transaction ID")
        except Exception:
            pass  # field not present for this payment mode — skip

        if debug:
            self.debug_modal_fields(modal, tag="after-fill")

        # ── 9. Click "Add Payment" ─────────────────────────────────────
        print("  Clicking 'Add Payment'...")
        add_payment_btn = modal.locator("button[title='Add Payment']")

        if add_payment_btn.is_disabled():
            print("  ⚠️  'Add Payment' still DISABLED — a field may be invalid.")
            _ss = os.path.join(_screenshots_dir, f"DISABLED_{invoice_no}.png")
            self.page.screenshot(path=_ss, full_page=False)
            print(f"  📸 {_ss} saved.")
            raise RuntimeError(
                f"'Add Payment' button disabled for {invoice_no} [{payment_mode}] — "
                "check the modal fields. Screenshot saved."
            )
        else:
            self._safe_mouse_click(add_payment_btn, label="Add Payment")
            self.page.wait_for_timeout(800)

        # ── 10. Screenshot before saving ──────────────────────────────
        _ss = os.path.join(_screenshots_dir, f"02_modal_{invoice_no}_{payment_mode}.png")
        self.page.screenshot(path=_ss, full_page=False)
        print(f"  📸 {_ss} saved.")

        # ── 11. Save inside modal ──────────────────────────────────────
        save_btn = modal.get_by_role("button", name="Save", exact=True)
        self._safe_mouse_click(save_btn, label="modal Save")
        print("  ✅ Modal Save clicked.")

        # ── 12. Wait for modal to close ───────────────────────────────
        modal.wait_for(state="hidden", timeout=15000)
        self._wait_mask_gone()
        self.page.wait_for_timeout(1000)

        _ss = os.path.join(_screenshots_dir, f"03_done_{invoice_no}_{payment_mode}.png")
        self.page.screenshot(path=_ss, full_page=True)
        print(f"  ✅ Done → {_ss} saved.")

    # ── FINAL PAGE SAVE ────────────────────────────────────────────────
    def save_page(self):
        """Commit all rows to the portal."""
        print("\nClicking final page Save...")
        self._wait_mask_gone()
        save_btn = self.page.get_by_role("button", name="Save", exact=True).last
        save_btn.scroll_into_view_if_needed()
        save_btn.click()
        self.page.wait_for_timeout(3000)
        self._wait_mask_gone()
        _ss = os.path.join(_screenshots_dir, "04_final_save.png")
        self.page.screenshot(path=_ss, full_page=True)
        print(f"✅ Final Save complete → {_ss} saved.")

    def close(self):
        print("Closing browser...")
        if self.browser:
            self.browser.close()
        if self.playwright:
            self.playwright.stop()


# ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Starting Portal Engine...")
    config = load_config()
    engine = RetailerPortalEngine(config)

    try:
        engine.launch()
        engine.login()

        # Read sheet — date auto-detected from first data row
        csv_path = config.get("sheet_path", r"D:\TATA_SHIPMENTS.xlsx")
        print(f"Reading sheet: {csv_path}")
        rows, run_date = load_rows_from_csv(csv_path)

        if not rows:
            print("⚠️  No processable rows found in sheet.")
            input("\nPress Enter to close...")
        else:
            print(f"\nDate from sheet  : {run_date}")
            print(f"Entries to process: {len(rows)}")
            for r in rows:
                print(f"  {r['invoice_no']}  {r['payment_mode']:8}  Rs.{r['amount']}")

            # Navigate using date from the sheet — no manual config needed
            engine.navigate_to_settlement_page(run_date)

            # Process each payment entry
            for row in rows:
                engine.fill_one_row(
                    invoice_no   = row["invoice_no"],
                    payment_mode = row["payment_mode"],
                    cheque_no    = row.get("cheque_no", ""),
                    amount       = row["amount"],
                    date         = row["date"],
                )

            print("\n✅ All entries processed.")

            # Pause before final page Save so you can review the full table
            print(f"\n{'='*55}")
            print("  ⏸  All modal entries saved.")
            print("  🔍 Review the table in the browser before final Save.")
            print("  ▶  Press Enter to click the final page Save button...")
            print(f"{'='*55}")
            input()

            engine.save_page()

        input("\nPress Enter to close browser...")

    except Exception as e:
        import traceback
        print("\n❌ Error:", e)
        traceback.print_exc()
        try:
            _ss = os.path.join(_screenshots_dir, "ERROR_state.png")
            engine.page.screenshot(path=_ss, full_page=True)
            print(f"📸 {_ss} saved.")
        except Exception:
            pass
        input("\nBrowser kept open. Press Enter to close...")

    finally:
        engine.close()